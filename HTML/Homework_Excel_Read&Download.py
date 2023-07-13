from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import gdown
import os
from concurrent.futures import ThreadPoolExecutor   # 加入 concurrent.futures 內建函式庫

# 設定---------------------------------------------------

path_ori = "E:/Desktop/6th_cadre/"                #輸入檔案要存到的資料夾絕對路徑
num_of_read = 1                                 #從第幾行excel開始讀
excel_path = 'E:/Desktop/6th_cadre/cadre.xlsx'      #excel檔的絕對路徑

# -------------------------------------------------------
wb = load_workbook(excel_path)
group = "設計" #input("請輸入組別:(設計、程控、機構、公關)")
path = ""
ws = wb.active
downloadlist = []
num_of_res = ws.max_row +1

# 編輯下載函式
def download(inf):
    code = inf[0]
    code = code.split("id=")
    code = code[1]
    if code != None:
        code = f"https://drive.google.com/u/1/uc?id={code}&export=download&confirm=t&uuid=69bb284e-baa9-4d04-b3dd-9c1d5fe2ab71&at=AHV7M3c1dcFHc-HsrO-MrJDXZj1B:1669515851648"
        # code = f"https://drive.google.com/u/1/uc?id={code}&export=download"
        os.chdir(inf[1])
        print(inf[1])
        gdown.download(code)
        print ("downloading")

#主程式
for raw in range(num_of_read, num_of_res):
    print(raw)
    downloadlist = []
    try:
        classname = int(ws['D'+str(raw)].value)
    except:
        classname = ws['D'+str(raw)].value
    path = str(classname) + str(ws['E'+str(raw)].value)

    try:
        os.mkdir(path_ori+path)
    except:
        pass

    if str(ws["F"+str(raw)].value) == group:
        i = 7
        for col in range(1, 11):
            totorial = get_column_letter(i)
            file_ch = get_column_letter(i+1)
            file_name = ws[totorial + str(raw)].value
            dwfile = str(ws[file_ch + str(raw)].value)
            url = dwfile.split(",")
            num = 0
            for f in url:
                if file_name == None:
                    pass
                else:
                    output = path_ori + path  #+ "/" + file_name + f"({str(num)})"+ "."+ file_type
                    con = f
                    downloadlist.append([con, output, path])
                    num += 1
            i +=3

        executor = ThreadPoolExecutor()  # 建立非同步的多執行緒的啟動器
        with ThreadPoolExecutor() as executor:
            executor.map(download, downloadlist)